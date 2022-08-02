from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt    # 匯入 matplotlib 程式庫
import os
import sys

if sys.platform.startswith("linux"):  # could be "linux", "linux2", "linux3", ...
    print("linux")              # linux
elif sys.platform == "darwin":  # MAC OS X
    plt.rcParams['font.sans-serif'] = 'Arial Unicode MS'
    # //注意這裡用的不是'SimHei'
    plt.rcParams['axes.unicode_minus'] = False
elif sys.platform == "win32":   # Windows (either 32-bit or 64-bit)
    plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei']
    plt.rcParams['axes.unicode_minus'] = False  # 步驟二（解決座標軸負數的負號顯示問題）
# ----------------------------------------------------------------------------------------
# 檔案變數

file_path="健保特約醫事機構-區域醫院_健保特約醫事機構_區域醫院 .xlsx"
wb = load_workbook(file_path)
sheet = wb.active
# 版本控制測試用新增段落

"""
# 修改工作表的資料，方法有二種，by直接定位 and 相對定位
sheet['A1'] = 87                  # 設定資料   A1
sheet.cell(row=1, column=2).value = 'OpenPyxl Tutorial' # 設定資料 B1
wb.save("sample_file.xlsx")
"""
# print(sheet.max_column)
# print(sheet.max_row) # 83，代表1~83，其中1為標頭

# TODO 1.用業務組區分管轄的區域醫院

def add_value_label(x_list,y_list):
    if type(x_list) !="list":
        x_list=list(x_list)
    if type(y_list) !="list":
        y_list = list(y_list)
    for i in range(len(x_list)):
        plt.text(x_list[i],y_list[i],y_list[i], ha="center")

def bar_hospital_nums_filter_by_supervise(sheet):
    supervise_team={}
    for i in range(2, sheet.max_row+1):
        team=sheet.cell(i,6).value          # 定義業務組變數
        if team not in supervise_team:
            supervise_team[team] = 1
        else:
            supervise_team[team] += 1

    plt.bar(supervise_team.keys(),supervise_team.values())
    add_value_label(supervise_team.keys(),supervise_team.values())
    plt.title("區域醫院相關統計")
    plt.xlabel("各區業務組")
    plt.ylim(0,30)
    plt.ylabel("\n權管地區\n醫院數",rotation=360,loc="center",labelpad=30)


bar_hospital_nums_filter_by_supervise(sheet)
plt.show()
# TODO 2.建立多和一的圖表(預定3合1)

def all_in_one(sheet):

    plt.subplot(2,2,1)
    bar_hospital_nums_filter_by_supervise(sheet)
    # 抓取縣市所含醫院數資料
    city_or_county_hospital_number = {}
    for i in range(2, sheet.max_row + 1):
        city_or_county = sheet.cell(i, 5).value[:3]  # 定義縣市別
        if city_or_county not in city_or_county_hospital_number:
            city_or_county_hospital_number[city_or_county] = 1
        else:
            city_or_county_hospital_number[city_or_county] += 1
    # 抓取醫院提供治療項目
    cure_items = {}
    for i in range(2, sheet.max_row + 1):
        cures = sheet.cell(i, 8).value.split(",")  # 門診項目
        for item in cures:
            if item not in cure_items:
                cure_items[item] = 1
            else:
                cure_items[item] += 1

    plt.subplot(2, 2, 2)
    plt.bar(city_or_county_hospital_number.keys(),
            city_or_county_hospital_number.values(),)
    plt.xticks(rotation=45)
    plt.ylabel("醫院數",rotation=360,loc="center",labelpad=30)
    plt.title("區域醫院縣市分布圖")
    add_value_label(city_or_county_hospital_number.keys(),
                    city_or_county_hospital_number.values())

    # 四癌篩檢
    four_types_cancer_check = {
        "婦女乳房檢查": cure_items["婦女乳房檢查"],
        "口腔黏膜檢查": cure_items['口腔黏膜檢查'],
        "婦女子宮頸抹片檢查": cure_items['婦女子宮頸抹片檢查'],
        "定量免疫法糞便潛血檢查": cure_items["定量免疫法糞便潛血檢查"]
    }

    non_check_hospital = [82 - int(four_types_cancer_check["婦女乳房檢查"]),
                          82 - int(four_types_cancer_check["口腔黏膜檢查"]),
                          82 - int(four_types_cancer_check["婦女子宮頸抹片檢查"]),
                          82 - int(four_types_cancer_check["定量免疫法糞便潛血檢查"])]
    check_hospital = [int(i) for i in four_types_cancer_check.values()]

    plt.subplot(2, 2, 3)
    plt.bar(four_types_cancer_check.keys(), check_hospital, width=0.4, label="提供檢查醫院數")
    plt.bar(four_types_cancer_check.keys(), non_check_hospital, bottom=check_hospital, label="無提供檢查醫院數", width=0.4)
    plt.ylabel("檢查項目",rotation=360,loc="center",labelpad=30)
    # 使用loc和labelpad可分別調整label的位置
    plt.xticks(rotation=30)
    plt.title("區域醫院四癌篩檢提供比例")
    plt.ylim(0,120)
    plt.legend(prop={'size': 8}, loc="upper right")


    sb=plt.subplot(2,2,4)
    district_hospital_cure_diversity={}
    for i in range(2, sheet.max_row + 1):
        team=sheet.cell(i,6).value
        if team not in district_hospital_cure_diversity:
            district_hospital_cure_diversity[team] ={
                sheet.cell(i, 2).value:len(sheet.cell(i, 8).value.split(","))
            }
        else:
            if sheet.cell(i, 2).value not in district_hospital_cure_diversity[team]:
                district_hospital_cure_diversity[team][sheet.cell(i, 2).value] = len(sheet.cell(i, 8).value.split(","))
    x=list(district_hospital_cure_diversity.keys())
    y=list(district_hospital_cure_diversity.values())
    count = 0
    district = 0
    for i in y:
        x_label=[num for num in range(count, count+len(i))]
        plt.scatter(x_label,i.values(),label=x[district])
        count += len(i)
        district += 1
    plt.title("區域醫院醫療服務多樣性(以各區業務組進行分區)")
    sb.get_xaxis().set_visible(False)
    plt.legend(prop={'size': 8},loc=(-0.3,0.15))
# plt.scatter(x,)
# print(y)




# plot2=plt.subplot(2,2,2)
# plt.pie(city_or_county_hospital_number.values(),
#         labels=city_or_county_hospital_number.keys(),
#         radius=50,
#         autopct='%1.1f%%',
#         wedgeprops={"linewidth": 1,
#                     "edgecolor": "white"},
#         frame=True)

all_in_one(sheet)
plt.tight_layout()
plt.show()